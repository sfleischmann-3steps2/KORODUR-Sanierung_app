#!/usr/bin/env python3
"""
Erstellt eine Excel-Vorlage zum Erfassen neuer Referenzen für die KORODUR Sanierungs-App.

Enthält 3 Tabellenblätter:
1. Referenzen (Stammdaten + Qualifizierung)
2. Produkt-Qualifizierung (Eignungsmatrix)
3. Ausfüllhilfe (Dropdown-Werte + Erklärungen)

Usage:
    python scripts/create_referenz_template.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# --- Styles ---
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="002D59", end_color="002D59", fill_type="solid")
SECTION_FILL = PatternFill(start_color="009EE3", end_color="009EE3", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
KO_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
KO_FONT = Font(name="Calibri", bold=True, size=11, color="CC0000")
EXAMPLE_FONT = Font(name="Calibri", italic=True, size=10, color="888888")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

# --- Produkte (aktuell in der App) ---
PRODUKTE = [
    "NEODUR HE 60 rapid",
    "NEODUR HE 65 Plus",
    "NEODUR HE 65",
    "NEODUR Level",
    "KORODUR FSCem Screed",
    "KORODUR HB 5 rapid",
    "KORODUR PC",
    "Rapid Set CEMENT ALL",
    "Rapid Set MORTAR MIX",
    "Rapid Set MORTAR MIX DUR",
    "Rapid Set CONCRETE MIX",
    "KOROCRETE Schnellbeton",
    "ASPHALT REPAIR MIX",
    "MICROTOP TW",
    "TRU Self-Leveling",
    "KOROCURE",
    "KOROMINERAL CURE",
    "KOROTEX",
]


def style_header(ws, row, num_cols, fill=HEADER_FILL, font=HEADER_FONT):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER


def auto_width(ws, min_width=12, max_width=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(cell.value or "")) for cell in col_cells),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def create_referenzen_sheet(wb):
    """Blatt 1: Referenz-Stammdaten + Qualifizierung"""
    ws = wb.active
    ws.title = "Referenzen"

    # Spalten-Definition: (Header, Breite, Beispielwert, Gruppe)
    columns = [
        # --- Stammdaten ---
        ("ID / Slug\n(eindeutig, Kleinbuchstaben,\nBindestrich)", 22, "muster-firma-halle", "Stammdaten"),
        ("Titel", 28, "Produktionshalle Muster GmbH", "Stammdaten"),
        ("Untertitel", 28, "Schnell saniert – voll belastbar", "Stammdaten"),
        ("Ort", 18, "München", "Stammdaten"),
        ("Land", 14, "Deutschland", "Stammdaten"),
        ("Fläche", 14, "500 m²", "Stammdaten"),
        ("Produkte\n(kommagetrennt)", 30, "NEODUR HE 60 rapid, KORODUR HB 5 rapid", "Stammdaten"),
        ("Herausforderungen\n(je Zeile eine,\ngetrennt durch |)", 45, "Hohe mechanische Belastung | Kurzes Zeitfenster | Alter Betonuntergrund", "Stammdaten"),
        ("Lösung\n(Fließtext)", 50, "Sanierung mit NEODUR HE 60 rapid auf Haftbrücke KORODUR HB 5 rapid. Nach 24 h voll belastbar.", "Stammdaten"),
        ("Vorteile\n(je Zeile eine,\ngetrennt durch |)", 45, "Kurze Ausfallzeit | Hohe Verschleißfestigkeit | Wirtschaftliche Lösung", "Stammdaten"),
        ("Bild-Dateiname\n(.jpg)", 20, "muster-firma.jpg", "Stammdaten"),
        ("Bild-Alttext", 30, "Sanierte Produktionshalle bei Muster GmbH in München", "Stammdaten"),

        # --- Einordnung Portfolio ---
        ("Kategorie", 18, "industrieboden", "Portfolio"),
        ("Unterkategorie", 18, "schwerlast", "Portfolio"),
        ("Anwendungsbereich", 20, "produktionshalle", "Portfolio"),

        # --- Qualifizierung Lösungsfinder ---
        ("Maßnahme", 22, "grossflaechige-sanierung", "Qualifizierung"),
        ("Zeitrahmen\n(K.O.)", 18, "ueber-nacht", "Qualifizierung K.O."),
        ("Umgebung\n(K.O.)", 18, "innen", "Qualifizierung K.O."),
        ("Belastungen\n(kommagetrennt)", 28, "schwerlast, rollende-lasten", "Qualifizierung"),
        ("Schadensbild\n(kommagetrennt)", 28, "abrieb-verschleiss, risse-ausbrueche", "Qualifizierung"),
        ("Sonderanforderungen\n(kommagetrennt)", 28, "chemikalien", "Qualifizierung"),

        # --- Übersetzungen ---
        ("Titel EN", 28, "Production Hall Muster GmbH", "Übersetzung"),
        ("Untertitel EN", 28, "Quick renovation – fully operational", "Übersetzung"),
        ("Titel FR", 28, "Hall de production Muster GmbH", "Übersetzung"),
        ("Untertitel FR", 28, "Rénovation rapide – pleinement opérationnel", "Übersetzung"),
        ("Titel PL", 28, "Hala produkcyjna Muster GmbH", "Übersetzung"),
        ("Untertitel PL", 28, "Szybka renowacja – w pełni operacyjna", "Übersetzung"),
    ]

    # Gruppen-Header (Zeile 1)
    group_ranges = {}
    for i, (_, _, _, group) in enumerate(columns, 1):
        base_group = group.replace(" K.O.", "")
        if base_group not in group_ranges:
            group_ranges[base_group] = [i, i]
        else:
            group_ranges[base_group][1] = i

    ws.row_dimensions[1].height = 22
    for group_name, (start, end) in group_ranges.items():
        fill = SECTION_FILL
        font = SECTION_FONT
        if group_name == "Qualifizierung":
            pass  # normal section color
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start, value=group_name)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Spalten-Header (Zeile 2)
    ws.row_dimensions[2].height = 55
    for i, (header, width, _, group) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=i, value=header)
        if "K.O." in group:
            cell.font = KO_FONT
            cell.fill = KO_FILL
        else:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width

    # Beispielzeile (Zeile 3)
    ws.row_dimensions[3].height = 60
    for i, (_, _, example, _) in enumerate(columns, 1):
        cell = ws.cell(row=3, column=i, value=example)
        cell.font = EXAMPLE_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

    # Leere Zeilen für Eingabe
    for row in range(4, 54):
        for col in range(1, len(columns) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = WRAP

    # Data Validations
    # Kategorie
    dv_kat = DataValidation(
        type="list",
        formula1='"industrieboden,industriebau,infrastruktur"',
        allow_blank=True,
    )
    dv_kat.error = "Bitte wählen: industrieboden, industriebau, infrastruktur"
    ws.add_data_validation(dv_kat)
    dv_kat.add(f"M3:M53")

    # Unterkategorie
    dv_ukat = DataValidation(
        type="list",
        formula1='"schwerlast,duennschicht,schnelle-reparaturen,fugen,verkehr,wasser"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_ukat)
    dv_ukat.add(f"N3:N53")

    # Anwendungsbereich
    dv_ab = DataValidation(
        type="list",
        formula1='"produktionshalle,lager,werkstatt,zufahrt,parkflaeche,bruecke,hafen,sonstiges"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_ab)
    dv_ab.add(f"O3:O53")

    # Maßnahme
    dv_mass = DataValidation(
        type="list",
        formula1='"punktuelle-reparatur,grossflaechige-sanierung,ausgleich-nivellierung,schutzbeschichtung"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_mass)
    dv_mass.add(f"P3:P53")

    # Zeitrahmen
    dv_zeit = DataValidation(
        type="list",
        formula1='"sofort,ueber-nacht,wenige-tage,geplant"',
        allow_blank=True,
    )
    dv_zeit.error = "K.O.-Kriterium! Bitte wählen: sofort / ueber-nacht / wenige-tage / geplant"
    ws.add_data_validation(dv_zeit)
    dv_zeit.add(f"Q3:Q53")

    # Umgebung
    dv_umg = DataValidation(
        type="list",
        formula1='"innen,aussen,frost-tausalz,nassbereich"',
        allow_blank=True,
    )
    dv_umg.error = "K.O.-Kriterium! Bitte wählen: innen / aussen / frost-tausalz / nassbereich"
    ws.add_data_validation(dv_umg)
    dv_umg.add(f"R3:R53")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:AA53"


def create_produkt_qualifizierung_sheet(wb):
    """Blatt 2: Produkt-Eignungsmatrix"""
    ws = wb.create_sheet("Produkt-Qualifizierung")

    headers = [
        "Produkt",
        # Maßnahme
        "Punktuelle\nReparatur", "Großflächige\nSanierung", "Ausgleich /\nNivellierung", "Schutzbeschichtung",
        # Zeitrahmen (K.O.)
        "Sperrzeitklasse\n(K.O.)",
        # Umgebung (K.O.)
        "Innen", "Außen", "Frost-/\nTausalz", "Nassbereich",
        # Belastung (3-stufig)
        "Schwerlast", "Rollende\nLasten", "Punktlasten", "Leichte\nNutzung",
        # Schadensbild
        "Risse /\nAusbrüche", "Abrieb /\nVerschleiß", "Hohlstellen", "Beschichtungs-\nschäden", "Ebenheits-\nprobleme", "Fugen\ndefekt",
        # Sonderanforderungen
        "Chemikalien-\nbeständig", "Rutsch-\nhemmung", "Normkonform\n(welche?)", "Design-\nanspruch", "Maschinelle\nVerarbeitung",
        # Kommentar
        "Kommentar /\nAnmerkung",
    ]

    # Gruppen-Header (Zeile 1)
    groups = [
        ("", 1, 1),
        ("Maßnahme", 2, 5),
        ("Sperrzeit", 6, 6),
        ("Umgebung (K.O.)", 7, 10),
        ("Belastung (0/1/2)", 11, 14),
        ("Schadensbild", 15, 20),
        ("Sonderanforderungen", 21, 25),
        ("", 26, 26),
    ]

    ws.row_dimensions[1].height = 22
    for group_name, start, end in groups:
        if not group_name:
            continue
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start, value=group_name)
        if "K.O." in group_name:
            cell.font = KO_FONT
            cell.fill = KO_FILL
        else:
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers (Zeile 2)
    ws.row_dimensions[2].height = 55
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=header)
        # K.O. columns
        if i == 6 or i in (7, 8, 9, 10):
            cell.font = KO_FONT
            cell.fill = KO_FILL
        else:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN_BORDER

    # Produkte eintragen
    for row_idx, produkt in enumerate(PRODUKTE, 3):
        ws.cell(row=row_idx, column=1, value=produkt).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        for col in range(2, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Data Validations
    last_row = len(PRODUKTE) + 2

    # Maßnahme: ja/nein
    dv_jn = DataValidation(type="list", formula1='"ja,nein"', allow_blank=True)
    ws.add_data_validation(dv_jn)
    for col in range(2, 6):  # B-E: Maßnahme
        dv_jn.add(f"{get_column_letter(col)}3:{get_column_letter(col)}{last_row}")

    # Sperrzeitklasse
    dv_sz = DataValidation(
        type="list",
        formula1='"sofort,ueber-nacht,wenige-tage,geplant,—"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_sz)
    dv_sz.add(f"F3:F{last_row}")

    # Umgebung: ja/nein
    dv_umg = DataValidation(type="list", formula1='"ja,nein"', allow_blank=True)
    ws.add_data_validation(dv_umg)
    for col in range(7, 11):  # G-J: Umgebung
        dv_umg.add(f"{get_column_letter(col)}3:{get_column_letter(col)}{last_row}")

    # Belastung: 0/1/2
    dv_012 = DataValidation(
        type="list",
        formula1='"0,1,2"',
        allow_blank=True,
    )
    dv_012.error = "0 = nicht geeignet, 1 = geeignet, 2 = ideal"
    ws.add_data_validation(dv_012)
    for col in range(11, 15):  # K-N: Belastung
        dv_012.add(f"{get_column_letter(col)}3:{get_column_letter(col)}{last_row}")

    # Schadensbild: ja/nein
    dv_sd = DataValidation(type="list", formula1='"ja,nein"', allow_blank=True)
    ws.add_data_validation(dv_sd)
    for col in range(15, 21):  # O-T: Schadensbild
        dv_sd.add(f"{get_column_letter(col)}3:{get_column_letter(col)}{last_row}")

    # Sonderanforderungen: ja/nein (21-22, 24-25), Freitext (23)
    dv_sa = DataValidation(type="list", formula1='"ja,nein"', allow_blank=True)
    ws.add_data_validation(dv_sa)
    for col in [21, 22, 24, 25]:
        dv_sa.add(f"{get_column_letter(col)}3:{get_column_letter(col)}{last_row}")

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 28
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions[get_column_letter(len(headers))].width = 30  # Kommentar

    ws.freeze_panes = "B3"


def create_ausfuellhilfe_sheet(wb):
    """Blatt 3: Erklärungen und erlaubte Werte"""
    ws = wb.create_sheet("Ausfüllhilfe")

    sections = [
        ("Kategorien (Portfolio)", [
            ("industrieboden", "Industrieböden: Produktionshallen, Lager, Werkstätten"),
            ("industriebau", "Industriebau: Fugen, konstruktive Bauteile"),
            ("infrastruktur", "Infrastruktur: Brücken, Verkehr, Wasser"),
        ]),
        ("Unterkategorien", [
            ("schwerlast", "Böden mit hoher mechanischer Belastung (Stapler, LKW)"),
            ("duennschicht", "Dünne Beschichtungen / Dünnestrich (4–15 mm)"),
            ("schnelle-reparaturen", "Punktuelle Reparaturen mit kurzer Sperrzeit"),
            ("fugen", "Fugen, Profile, konstruktive Übergänge"),
            ("verkehr", "Brücken, Parkflächen, Zufahrten"),
            ("wasser", "Trinkwasserbehälter, Wasserbauwerke"),
        ]),
        ("Anwendungsbereiche", [
            ("produktionshalle", "Produktion, Fertigung, Montage"),
            ("lager", "Lager, Logistik, Distribution"),
            ("werkstatt", "KFZ-Werkstatt, Prüfstände"),
            ("zufahrt", "Einfahrt, Umfahrt, Tankstelle"),
            ("parkflaeche", "Parkhaus, Parkplatz"),
            ("bruecke", "Brücken, Überführungen"),
            ("hafen", "Hafenflächen, Container-Terminals"),
            ("sonstiges", "Alles andere (Trinkwasser, Retail, Helipad, ...)"),
        ]),
        ("Maßnahme (Schritt 1 im Wizard)", [
            ("punktuelle-reparatur", "Einzelne Schadstellen reparieren: Löcher, Risse, Ausbrüche, Fugen"),
            ("grossflaechige-sanierung", "Ganze Bereiche komplett erneuern"),
            ("ausgleich-nivellierung", "Ebenheit herstellen, Höhenunterschiede ausgleichen"),
            ("schutzbeschichtung", "Oberfläche langfristig schützen (Trinkwasser, Verschleiß, Chemie)"),
        ]),
        ("Zeitrahmen / Sperrzeit (K.O.-Kriterium)", [
            ("sofort", "< 2 Stunden bis zur Nutzung (Notfall, laufender Betrieb)"),
            ("ueber-nacht", "< 24 Stunden (Nachtschicht / Wochenende)"),
            ("wenige-tage", "1–5 Tage (geplante Betriebspause)"),
            ("geplant", "> 5 Tage (reguläre Baumaßnahme, Zeit kein Engpass)"),
        ]),
        ("Umgebung (K.O.-Kriterium)", [
            ("innen", "Innenbereich, temperiert, keine Witterung"),
            ("aussen", "Außenbereich, witterungsexponiert, aber kein extremer Frost"),
            ("frost-tausalz", "Außenbereich MIT Frost-/Tausalzbelastung"),
            ("nassbereich", "Trinkwasser / Nassbereich (DVGW / WHG nötig)"),
        ]),
        ("Belastungen", [
            ("schwerlast", "Stapler, LKW, Container, Kettenfahrzeuge (> 5 t Achslast)"),
            ("rollende-lasten", "Hubwagen, Flurförderzeuge, PKW"),
            ("punktlasten", "Regale, Maschinen, Stützen (statische Einzellasten)"),
            ("leichte-nutzung", "Fußgänger, leichte Wagen, Retail"),
        ]),
        ("Schadensbild", [
            ("risse-ausbrueche", "Strukturelle Risse, Kantenbrüche, Ausbrüche"),
            ("abrieb-verschleiss", "Oberflächenabtrag, Staubbildung, rauer Boden"),
            ("hohlstellen", "Keine Haftung zum Untergrund, hohler Klang"),
            ("beschichtungsschaeden", "Alte Beschichtung defekt, blättert, löst sich"),
            ("ebenheitsprobleme", "Unebenheiten, Höhenversatz, Stolperkanten"),
            ("fugen-defekt", "Fugenprofile verschlissen, ausgebrochen, nicht mehr tragfähig"),
        ]),
        ("Sonderanforderungen", [
            ("chemikalien", "Beständigkeit gegen Öle, Säuren, Reinigungsmittel"),
            ("rutschhemmung", "Rutschhemmende Oberfläche (R-Klassen)"),
            ("normkonform", "Bestimmte Normen zwingend (DIN EN 1504, DVGW, WHG, ...)"),
            ("designanspruch", "Optik wichtig: Sichtestrich, Betonoptik, Farbkonzept"),
            ("maschinelle-verarbeitung", "Silosystem, Pumptechnik für große Flächen"),
        ]),
        ("Belastungs-Eignungsstufen (nur Produkt-Qualifizierung)", [
            ("2 = Ideal", "Dafür gemacht, vielfach bewährt, erste Wahl"),
            ("1 = Geeignet", "Funktioniert, aber nicht optimiert dafür"),
            ("0 = Nicht geeignet", "Produkt kann diese Belastung nicht leisten"),
        ]),
        ("Produkte (aktuell in der App)", [
            (p, "") for p in PRODUKTE
        ]),
    ]

    row = 1
    for section_title, items in sections:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value=section_title)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        for col in range(1, 4):
            ws.cell(row=row, column=col).fill = SECTION_FILL
        row += 1

        # Column headers
        ws.cell(row=row, column=1, value="Wert").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Beschreibung").font = Font(bold=True)
        row += 1

        for value, description in items:
            ws.cell(row=row, column=1, value=value).font = Font(name="Consolas", size=10)
            ws.cell(row=row, column=2, value=description)
            row += 1

        row += 1  # Leerzeile

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 65


def main():
    wb = Workbook()

    create_referenzen_sheet(wb)
    create_produkt_qualifizierung_sheet(wb)
    create_ausfuellhilfe_sheet(wb)

    output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "docs")
    output_path = os.path.join(output_dir, "KORODUR_Referenz_Vorlage.xlsx")
    wb.save(output_path)
    print(f"Excel-Vorlage erstellt: {output_path}")


if __name__ == "__main__":
    main()
